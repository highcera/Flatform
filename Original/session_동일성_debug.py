import openpyxl as op
import numpy as np

#path는 대상이 되는 엑셀파일의 절대 경로
def corretion(path : str):
    #openpyxl 워크북 정의
    wb = op.load_workbook(path)
    #워크북의 '동일성' 시트를 객체로 정의
    ws = wb["I_AI9MP1U_2"]

    #해당 시트의 마지막 열, 마지막 행 
    column_max = ws.max_column
    row_max = ws.max_row
    no_of_comb = int(input("Set 수는? : "))
    print(column_max, row_max, "\n")

    # kai = np.zeros((no_of_comb, 3))
    # cond = np.array((no_of_comb, 3))
    
    #열마다 행을 for loop문 진행
    for col_num in range(2, 3):
        for row_num in range(1, row_max+1):
            tempstr = str(ws.cell(row = row_num, column = col_num).value)
            
            cond = []
            kai = []

            if tempstr == '카이-제곱':
                for y in range(0, 3):
                    cond.append(ws.cell(row = row_num-4, column = col_num+y+2).value)
                    # cond = ws.cell(row = row_num-4, column = col_num+y+2).value
                # print(cond)

                aaa = ws.cell(row = row_num+1, column = col_num).value
                if aaa == None:
                    for y in range(0, 3):
                        kai.append(ws.cell(row = row_num+1, column = col_num+y+1).value)
                        # kai = ws.cell(row = row_num+1, column = col_num+y+1).value
                    # print(kai)
                else:
                    for y in range(0, 3):     
                        kai.append(ws.cell(row = row_num+1, column = col_num+y).value)
                        # kai = ws.cell(row = row_num+1, column = col_num+y).value
                    # print(kai)

                total_info = cond + kai
                # ws.cell(row = row_num+2, column = 300, value = total_info)
                print(total_info)
            
        # 저장
        # wb.save(path)

if __name__ == "__main__":
    path = r"P:/Intel 자동화 검증/03W105MR3_AI9MP1U_3X4_2_session.xlsx"

    #함수 실행
    corretion(path)
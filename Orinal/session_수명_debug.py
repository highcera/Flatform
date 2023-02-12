import openpyxl as op
import numpy as np

#path는 대상이 되는 엑셀파일의 절대 경로
def corretion(path : str):
    #openpyxl 워크북 정의
    wb = op.load_workbook(path)
    #워크북의 활성화된 시트를 객체로 정의
    ws = wb["L_AI9MP1U_2"]
    # #시트 탭 색깔 변경(하늘색)
    # ws.sheet_properties.tabColor = '#008000'


    #해당 시트의 마지막 열, 마지막 행 
    column_max = ws.max_column
    row_max = ws.max_row
    no_of_comb = int(input("Set 수는? : "))
    print(column_max, row_max, no_of_comb)

    temp = np.zeros((no_of_comb, 6))
    volt = np.zeros((no_of_comb, 6))
    life = np.zeros((no_of_comb, 4))
    
    #열마다 행을 for loop문 진행
    for col_num in range(2, 4):
        for row_num in range(1, row_max+1):
            for x in range(0, no_of_comb):
                #tempstr : cell값이 문자열이 아닌 경우를 감안하여 str로 바꿔줌
                tempstr = str(ws.cell(row = row_num, column = col_num).value)
                tempstr2 = str(ws.cell(row = row_num+1, column = col_num).value)
#                 if (tempstr == 'Temp'+str(x+1) and tempstr2 == 'Vot'+str(x+1)):
#                     print (col_num, row_num, tempstr, tempstr2)
                       
                if tempstr == 'Temp'+str(x+1):
                    if tempstr2 == 'Volt'+str(x+1):
                        for y in range(0, 6):
                            temp[x][y] = ws.cell(row = row_num, column = col_num+y+1).value
                            volt[x][y] = ws.cell(row = row_num+1, column = col_num+y+1).value
                    elif tempstr2 == '0.001':   
                        for y in range(0, 4):
                            life[x][y] = ws.cell(row = row_num+1, column = col_num+y+3).value
               
             
                    
                                
    print("\n")                                              
    print(temp, "\n", volt, "\n", life)

    print("==============================    Extraction Result   =============================") 
    print(" Set No              Temp Accel_Low             Volt Accel_Low             Life_Low")
    for i in range(0, no_of_comb):
        print('Set'+str(i+1), temp[i][4], abs(volt[i][5]), life[i][2]/8750)
          

    #저장
#     wb.save("correction_result.xlsx")

if __name__ == "__main__":
#     path = input("화일경로는 ? \n Ex) r'P:\자동데이타생성\쎄션 데이터 추출.xlsx' : ")
    path = r"P:/Intel 자동화 검증/03W105MR3_AI9MP1U_3X4_2_session.xlsx"

    #함수 실행
    corretion(path)
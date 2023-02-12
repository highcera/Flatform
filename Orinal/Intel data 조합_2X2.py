import sys, os
import copy

import pandas as pd
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
import openpyxl as op
import xlwings as xw

import numpy as np
# import matplotlib as mpl

from itertools import combinations, product

# font_name = mpl.font_manager.FontProperties(fname='C:/Windows/Fonts/malgun.ttf').get_name()
# mpl.rc('font', family=font_name)
# mpl.rc('axes', unicode_minus = False)

MAIN_FILE_NAME = 'main_wnd_intel'
os.system('python -m PyQt5.uic.pyuic -x ' + MAIN_FILE_NAME + '.ui -o ' + MAIN_FILE_NAME + '.py')

from main_wnd_intel import Ui_MainWindow
from pandasModel_var import *

class Form(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()

        self.setupUi(self)
        self.actionOpen_2.triggered.connect(self.open_file)
       
#        self.btnApply.clicked.connect(self.apply_lib)
#        self.btnGraph.clicked.connect(self.draw_graph)
#        self.btnExcel.clicked.connect(self.make_excel)

        # self.prog_load.setHidden(True)
        # self.prog_disp.setHidden(True)
        # self.prog_conv.setHidden(True)
        # self.prog_filepre.setHidden(True)
        
        # self.btnGraph.setHidden(True)
        # self.fileLineEdit.setHidden(True)
        # self.btnExcel.setHidden(True)
        

        # self.df = pd.DataFrame()
        # self.lib = pd.DataFrame()
        # self.test = pd.DataFrame()


    def open_file(self):
        os.chdir("P:\Intel 자동화 검증")
        show_filter = "모든 파일(*.*);;엑셀파일(*.xls);; 엑셀파일(*.xlsx)"
        init_filter = "엑셀파일(*.xlsx)"
        opt = QFileDialog.Option()
        file_name, filter_type = QFileDialog.getOpenFileName(filter=show_filter, initialFilter=init_filter, options=opt)
     
        if file_name:
            self.setWindowTitle(file_name.split('/')[-1])

            # self.df = pd.read_excel(file_name, header = 1, usecols = "C:W", engine = 'openpyxl')

            bk = xw.Book(file_name)
            sh = bk.sheets(2) # 동일 화일 내 다른 sheet 사용 시 설정)
            print(sh.name)
            self.file_name.setText(file_name)
            self.part_name.setText(os.path.basename(file_name).split('.')[0].split(' ')[0])
            self.lot_no.setText(sh.name)

            # df1_des = bk.sheets(1).range('B3:M13').options(pd.DataFrame, index=False, header=False).value     # 가변 data에 대비해 수정 필요 (table... etc)
            df1_des = sh.range('B3:B13').options(pd.DataFrame, index=False, header=False, expand='right').value
            print(df1_des)
            # testNo = df1_des.iloc[0].to_list()
            # testDate = df1_des.iloc[1].to_list()
            # testFinTime = df1_des.iloc[2].to_list()
            # test_set_ID = df1_des.iloc[3].to_list()
            # equip = df1_des.iloc[4].to_list()
            # partNo = df1_des.iloc[5].to_list()
            # LotNo = df1_des.iloc[6].to_list()

            quantity = df1_des.iloc[7].to_list()
            temp = df1_des.iloc[8].to_list()
            volt = df1_des.iloc[9].to_list()
            time = df1_des.iloc[10].to_list()

            column = [str(int(temp[i])) + '-' + str(float(volt[i])) for i in range(len(temp))]
            print(temp, volt, quantity, time)  

            df1_data = sh.range('B14').options(pd.DataFrame, index=False, header=False, expand='table').value
            # df1_data = df1_data.round(1)
            print(df1_data)
            
            # print(df1_data)
            
            # 기종명, Lot no 추출
            # test_pn = list(set(partNo))
            # print(test_pn, test_pn[0])
            # self.part_name.setText(test_pn[0])  
            
            # test_ln = list(set(LotNo))
            # print(test_ln, test_ln[0])
            # self.lot_no.setText(test_ln[0])               
            
            # 시험온도, 시험전압 테이블 추출
            test_temp = list(set(temp))
            test_temp = [int(x) for x in test_temp if pd.isnull(x) == False]
            test_temp.sort()
            test_temp = [str(x) for x in test_temp if pd.isnull(x) == False]
            print(test_temp)
            
            test_volt = list(set(volt))
            test_volt = [float(x) for x in test_volt if pd.isnull(x) == False]
            test_volt.sort()
            test_volt = [str(x) for x in test_volt if pd.isnull(x) == False]
            print(test_volt)
            
            # 테스트 수량 추출
            test_quantity = list(set(quantity))
            test_quantity = [int(x) for x in test_quantity if pd.isnull(x) == False]    
            max_quantity = max(test_quantity)
            print(max_quantity)
            
            test_time = time
            print(test_time)
            
            self.prog_load.setText('Completed...')
            
            max_data = max_quantity
            
            i = len(test_temp)
            j = len(test_volt)
            k = max_data

            data_array= np.zeros((i, j, k))
            empty_cond = np.ones((i, j), dtype=np.int8)  

            count=0
            for x in range(len(test_temp)):
                for y in range(len(test_volt)):
                    data_array[x][y] = df1_data.loc[0:max_quantity-1, count]
                    data_array[x][y] = sorted(data_array[x][y])
                    if data_array[x][y][max_quantity-1]==0:
                        empty_cond[x][y] = 0
                    count += 1   
            
            print(data_array)

            testcond = pd.DataFrame(empty_cond, columns = test_volt, index = test_temp)
            
            self.testcond = pd.DataFrame(empty_cond, columns = test_volt, index = test_temp)
            self.testcond = self.testcond.fillna('V')
            self.make_tableConst(self.testcond)
             
            print(testcond)    

            # self.prog_load.setHidden(False)
            self.prog_load.setText('Data Loading completed...')
                   
            show_data = df1_data.copy()
            # show_data = show_data.round(2)
            show_data.columns = column
            new_index = range(1, max_data+1)
            show_data.index = new_index
            self.make_tableLib(show_data)
  
            # self.prog_disp.setHidden(False)
            self.prog_disp.setText('Completed...')


            # Intel 온도-전압 전체 조합 산출  
            temp_comb = list(combinations(list(range(len(test_temp))), 2))
            volt_comb = list(combinations(list(range(len(test_volt))), 2))
            
            unit_comb = list(product(temp_comb,volt_comb))

            no_of_comb = len(temp_comb)*len(volt_comb)
            comb_3_cond = np.zeros((no_of_comb*4, 3, max_quantity))
            ret = np.zeros((no_of_comb, 4, 3, 2))
            
            for i in range(0, no_of_comb):
                tot_set = list(product(unit_comb[i][0], unit_comb[i][1]))
                final_comb = list(combinations(tot_set, 3))
                # print(final_comb)
                ret[i] = np.array(final_comb)

            output_array = np.zeros((no_of_comb, 4, 3, max_quantity))
            
            for i in range(0, no_of_comb):
                for j in range(0,4):
                    for k in range(0,3):
                        output_array[i][j][k] = data_array[int(ret[i][j][k][0])][int(ret[i][j][k][1])]    
            print(output_array)

            title_array=[] 
            for i in range(0, no_of_comb):
                for j in range(0,4):
                    for k in range(0,3):
                        title_array.append(test_temp[int(ret[i][j][k][0])] + '-' + test_volt[int(ret[i][j][k][1])]) 
            print(title_array)   
            
            title_array2=[]
            for i in range(0, no_of_comb*4*3, 3):
                title_array2.append(title_array[i] + '/' + title_array[i+1] + '/' + title_array[i+2]) 
            print(title_array2)    
            
        
            reduction = np.reshape(output_array, (no_of_comb*4, 3, -1))
            
            remove_index = []
            for x in range(0,no_of_comb*4):    
                for y in range(0,3): 
                    if reduction[x][y][max_quantity-1] == 0:
                        remove_index.append(x)


            # self.prog_conv.setHidden(False)
            self.prog_conv.setText('Completed...')
            
            
            ### Exel 저장 준비
            ## 3개 조합 단위 데이터 출력 
            
            censoring = copy.deepcopy(reduction)
            
            wb1 = op.Workbook()
            
            ws1 = wb1.create_sheet("Full_comb_of_three_cond",0) #wb 객체를 통해 새로운 시트 생성
            ws1 = wb1["Full_comb_of_three_cond"] #결과 출력용 Sheet로 설정하기
            
            # title 출력
            i=0
            for x in range(0, no_of_comb*4):
                if x not in remove_index:
                    ws1.cell(row = 1, column = i*7+1).value = 'Set '+str(i+1) 
                    # 72 set 정보를 행(row=1)에 열을 바꾸면서 출력
                    i += 1

            i=1
            for x in range(0, no_of_comb*4):
                if x not in remove_index:
                    for y in range(0, 3):
                        # print(x*3+y, title_array[x*3+y])
                        ws1.cell(row = 2, column = i).value = title_array[x*3+y]  # 1행(row=1)에 열을 바꾸면서 출력
                        ws1.cell(row = 2, column = i+1).value = 'Censor'+str(x*3+y+1) 
                        i += 2
                    i += 1
                    
            i=3 #행
            j=1 #열
            for x in range(0, no_of_comb*4):    
                if x not in remove_index: 
                    for y in range(0,3): 
                        for z in  range(0,max_quantity):
                            if censoring[x][y][z] != 0: 
                                if censoring[x][y][z] == 999:                  
                                    target = title_array[x*3+y]
                                    censoring[x][y][z] = time[(column.index(target))]
                                    ws1.cell(row = i, column=j).value = censoring[x][y][z]  
                                    ws1.cell(row = i, column=j+1).value = 1 
                                    # print(x, '-', y, censoring[x][y][z])
                                else:
                                    ws1.cell(row = i, column=j).value = censoring[x][y][z]  
                                    ws1.cell(row = i, column=j+1).value = 0 
                                    # print(x, '-', y, censoring[x][y][z])  
                                i += 1  #행값 증가
                        i = 3
                        j += 2 # 열값 증가
                    j += 1 # 3열당 한칸 뜀
                    
            # ------------------------------------------------------------
            # 2. 미니탭 'Temp Volt Time censor' 형식 출력
            # ------------------------------------------------------------
            
            censoring = copy.deepcopy(reduction)

            ws2 = wb1.create_sheet("Minitab_data",1)  #wb 객체를 통해 새로운 시트 생성
            ws2 = wb1["Minitab_data"]                 #결과 출력용 Sheet로 설정하기
                        
            # title 출력
            i=0 
            for x in range(0, no_of_comb*4):
                if x not in remove_index:
                    ws2.cell(row = 1, column = i*5+1).value = 'Set '+str(i+1)+' : ' # 72 set 정보를 행(row=1)에 열을 바꾸면서 출력
                    ws2.cell(row = 1, column = i*5+2).value = title_array2[i]  
                    ws2.cell(row = 2, column = i*5+1).value = 'Temp'+str(i+1)
                    ws2.cell(row = 2, column = i*5+2).value = 'Volt'+str(i+1)
                    ws2.cell(row = 2, column = i*5+3).value = 'Time'+str(i+1)
                    ws2.cell(row = 2, column = i*5+4).value = 'Censor'+str(i+1)
                    i += 1
             
            title_array3 = np.reshape(title_array, (no_of_comb*4, -1))
          
            i=3
            j=1
            for x in range(0, no_of_comb*4):    
                if x not in remove_index:
                    for y in range(0, 3): 
                        for z in  range(0, max_quantity):
                            # print(censoring[x][y][z], reduction[x][y][z])
                            if censoring[x][y][z] != 0:
                                if int(censoring[x][y][z]) == 999:
                                    censor_temp = int(title_array3[x][y].split('-')[0])
                                    censor_volt = float(title_array3[x][y].split('-')[1])
                                    target = str(censor_temp)+'-'+str(censor_volt)
                                    censoring[x][y][z] = time[(column.index(target))]
                                    
                                    # print(int(title_array3[x][y].split('-')[0]), float(title_array3[x][y].split('-')[1]), censoring[x][y][z], "1")
                                    ws2.cell(row = i, column=j).value = int(title_array3[x][y].split('-')[0])
                                    ws2.cell(row = i, column=j+1).value = float(title_array3[x][y].split('-')[1])
                                    ws2.cell(row = i, column=j+2).value = censoring[x][y][z]          
                                    ws2.cell(row = i, column=j+3).value = 1
                                else:
                                    # print(int(title_array3[x][y].split('-')[0]), float(title_array3[x][y].split('-')[1]), censoring[x][y][z], "0")
                                    ws2.cell(row = i, column=j).value = int(title_array3[x][y].split('-')[0])
                                    ws2.cell(row = i, column=j+1).value = float(title_array3[x][y].split('-')[1])
                                    ws2.cell(row = i, column=j+2).value = censoring[x][y][z]   
                                    ws2.cell(row = i, column=j+3).value = 0
                                i += 1                    #행값 증가
                    i = 3
                    j += 5   
              
            wb1.save(os.path.splitext(file_name)[0]+'_'+sh.name+'_combi.xlsx') #해당 워크북(엑셀파일) 저장하기            
            
            
            
            
            
            
            
 #  --------------------------------------   기  존   -------------------------------------------------------------------------           
            # wb1 = op.Workbook() #새로운 워크북 객체 생성

            # ws1 = wb1.create_sheet("Full_comb_of_three_cond",0)     #wb 객체를 통해 새로운 시트 생성
            # ws1 = wb1["Full_comb_of_three_cond"]                    #결과 출력용 Sheet로 설정하기
            
            # # title 출력
            # i=1
            
            # for x in range(0, no_of_comb*4*3):
            #     ws1.cell(row = 1, column = i).value = title_array[x]  # 1행(row=1)에 열을 바꾸면서 출력
            #     if x%3 == 2:
            #         i += 2  # 3열에 구분열 1열씩 추가 
            #     else:
            #         i += 1  # 열값 증가
            
            # i=2
            # j=1
            # for x in range(0, no_of_comb*4):    
            #     for y in range(0,3): 
            #         for z in  range(0,max_quantity):
            #             ws1.cell(row = i, column=j).value = reduction[x][y][z]  #A열(Column=1)에 행을 바꾸면서 출력
            #             i += 1  #행값 증가
            #         i = 2
            #         j += 1 # 열값 증가
            #     i = 2
            #     j += 1
            
            
            # ## 미니탭 형식 'Temp1 Volt1 Time1 censor1' 출력
 
            # ws2 = wb1.create_sheet("Minitab_data",1)  #wb 객체를 통해 새로운 시트 생성
            # ws2 = wb1["Minitab_data"]                 #결과 출력용 Sheet로 설정하기
            
            # # title 출력
            # for x in range(0, no_of_comb*4):
            #     ws2.cell(row = 1, column = x*5+1).value = 'Set '+str(x+1)+' : ' # 72 set 정보를 행(row=1)에 열을 바꾸면서 출력
            #     ws2.cell(row = 1, column = x*5+2).value = title_array2[x]  
            #     ws2.cell(row = 2, column = x*5+1).value = 'Temp'+str(x+1)
            #     ws2.cell(row = 2, column = x*5+2).value = 'Volt'+str(x+1)
            #     ws2.cell(row = 2, column = x*5+3).value = 'Time'+str(x+1)
            #     ws2.cell(row = 2, column = x*5+4).value = 'Censor'+str(x+1)
             
            # title_array3 = np.reshape(title_array, (no_of_comb*4, -1))

            # i=3
            # j=1
            # for x in range(0, no_of_comb*4):    
            #     for y in range(0,3): 
            #         for z in  range(0,max_quantity):
            #             ws2.cell(row = i, column=j).value = int(title_array3[x][y].split('-')[0])
            #             ws2.cell(row = i, column=j+1).value = float(title_array3[x][y].split('-')[1])
            #             ws2.cell(row = i, column=j+2).value = reduction[x][y][z]                       # A열(Column=1)에 행을 바꾸면서 출력
            #             ws2.cell(row = i, column=j+3).value = 0                                        # Censoring 판단 로직 추가 필요
            #             i += 1                    #행값 증가
            #     i = 3
            #     j += 5   
           
  
# --------------------------------------------    비정형 포함    -----------------------------------------------------------------  
 
# ------------------------------------------------------------
# 결과 array data 화일 출력 - 1단계
# 72 X 3 = 216 Title (평가온도-평가전압) & data
# ------------------------------------------------------------

  
            
            
            # self.fileLineEdit.setHidden(False)
            # self.btnExcel.setHidden(False)
            
            # save_file = self.fileLineEdit.text()
            # if not save_file:
            #     QMessageBox.warning(self, 'warning', '파일 이름을 입력해주세요.')
            # else:
            #     sfile_name = r'P:\자동데이타생성\{}.xlsx'.format(save_file)
           
            # print(sfile_name)
            
            # wb1.save(r'P:\자동데이타생성\I2_3X4비정형_03W225.xlsx')   #해당 워크북(엑셀파일) 저장하기    
            
            # self.prog_filepre.setHidden(False)
            # self.prog_filepre.setText('Completed...')

    def make_tableLib(self, df):
        model = pandasModel(df)
        self.tableLib.setModel(model)
        self.tableLib.resizeColumnsToContents()
        self.tableLib.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableLib.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        print('make_tableLib 완료')


    def make_tableConst(self, df):
        self.tableConst.setHidden(False)
        model = pandasModel(df)
        self.tableConst.setModel(model)
        self.tableConst.resizeColumnsToContents()
        self.tableConst.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableConst.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        print('make_tableConst 완료')


#     def apply_lib(self):

    
#   
#         self.fin = self.test.T.reset_index(drop = True)
        
#       
#   
#         cap = self.CaplineEdit.text()



#         self.make_tableConst(self.cst)
    
#     def draw_graph(self):

#         x = [float(a) for a in self.fin['전압 (Vdc)'].values]
#         y = self.fin['용량변화율 %'].values


#         xs = np.linspace(x[0], x[-1], 1000)
#         f = interpolate.interp1d(x, y, kind = 'linear')

#         ys = f(xs)

#         if self.filtered_thr['DC-Bias 측정'][0] == '저주파':
#             xp = [float(a) for a in self.low_show['전압 (Vdc)'].values]
#             yp = self.low_show['용량변화율 %'].values
#         elif self.filtered_thr['DC-Bias 측정'][0] == '고주파':
#             xp = [float(a) for a in self.high_show['전압 (Vdc)'].values]
#             yp = self.high_show['용량변화율 %'].values

#         plt.scatter(xp,yp, color = 'red')
#         plt.plot(xs, ys)


#         plt.xlabel('전압 (V)')
#         plt.ylabel('용량변화율 %')

#         plt.grid(True)


#         plt.show()

    # def make_excel(self):
    #     new_file = self.fileLineEdit.text()
    #     if not new_file:
    #         QMessageBox.warning(self, 'warning', '파일 이름을 입력해주세요.')
    #     else:
    #         newfile = 'P:.\\..\\{}.xlsx'.format(new_file)
    #         # newfile = './{}.xlsx'.format(new_file)
           
    #         writer = pd.ExcelWriter(newfile)
            
    #         self.low.to_excel(writer, index=False, engine = 'openpyxl')
            
    #         writer.save()
            
    #         QMessageBox.information(self, '저장완료', '성공적으로 저장되었습니다.')
      

    # def my_exception_hook(exctype, value, traceback, self=None):
    #     # Print the error and traceback
    #     print(exctype, value, traceback)

    #     # Call the normal Exception hook after
    #     sys._excepthook(exctype, value, traceback)
    #     QMessageBox.warning(self, '오류', '올바르지 않은 접근입니다.')
    #     # sys.exit(1)

    # # Back up the reference to the exceptionhook
    # sys._excepthook = sys.excepthook

    # # Set the exception hook to our wrapping function
    # sys.excepthook = my_exception_hook

if __name__ == '__main__':
    app = QApplication(sys.argv)
    w = Form()
    w.show()
    sys.exit(app.exec_())